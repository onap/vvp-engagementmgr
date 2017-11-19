#
# ============LICENSE_START==========================================
# org.onap.vvp/engagementmgr
# ===================================================================
# Copyright © 2017 AT&T Intellectual Property. All rights reserved.
# ===================================================================
#
# Unless otherwise specified, all software contained herein is licensed
# under the Apache License, Version 2.0 (the “License”);
# you may not use this software except in compliance with the License.
# You may obtain a copy of the License at
#
#             http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
#
#
# Unless otherwise specified, all documentation contained herein is licensed
# under the Creative Commons License, Attribution 4.0 Intl. (the “License”);
# you may not use this documentation except in compliance with the License.
# You may obtain a copy of the License at
#
#             https://creativecommons.org/licenses/by/4.0/
#
# Unless required by applicable law or agreed to in writing, documentation
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# ============LICENSE_END============================================
#
# ECOMP is a trademark and service mark of AT&T Intellectual Property.
from django.db.models.expressions import F
from django.db.models.query_utils import Q
from django.http.response import HttpResponse
from django.utils.dateparse import parse_date
from engagementmanager.decorator.auth import auth
from engagementmanager.decorator.class_decorator import classDecorator
from engagementmanager.decorator.log_func_entry import logFuncEntry
from engagementmanager.git.git_manager import GitManager
from engagementmanager.models import Engagement, VF, VFC, Checklist, \
    EngagementStatus
from engagementmanager.rest.vvp_api_view import VvpApiView
from engagementmanager.serializers import VFModelSerializer, \
    EngagementStatusModelSerializer
from engagementmanager.service.authorization_service import Permissions
from engagementmanager.service import engagement_service as eng_svc
from engagementmanager.utils.constants import Roles, \
    RecentEngagementActionType, CheckListState, EngagementStage
from engagementmanager.utils.request_data_mgr import request_data_mgr
from engagementmanager.utils.validator import logEncoding
from rest_framework.response import Response
from rest_framework.status import HTTP_500_INTERNAL_SERVER_ERROR, \
    HTTP_400_BAD_REQUEST, HTTP_200_OK, HTTP_204_NO_CONTENT, \
    HTTP_401_UNAUTHORIZED, HTTP_202_ACCEPTED


@classDecorator([logFuncEntry])
class ExpandedEngByUser(VvpApiView):

    def post(self, request):
        data = request.data
        if ('stage' not in data or not data['stage']
            or 'keyword' not in data
            or 'offset' not in data or int(data['offset']) < 0
                or 'limit' not in data or not data['limit'] or
                (data['limit'] < 1)):
            msg = "GetExpandedEngByUser - get request: one of the parameters \
            is missing or invalid."
            self.logger.error(msg)
            msg = "Action was failed due to bad request."
            return Response(msg, status=HTTP_400_BAD_REQUEST)
        user = request_data_mgr.get_user()
        data = eng_svc.get_dashboard_expanded_engs(
            data['stage'], data['keyword'], data['offset'],
            data['limit'], user)
        return Response(data)


@classDecorator([logFuncEntry])
class ExportEngagements(VvpApiView):

    @auth(Permissions.export_engagments)
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.writer.write_only import WriteOnlyCell
        from openpyxl.styles import Font
        # import csv
        from django.utils.encoding import smart_str

        user = request_data_mgr.get_user()
        stageParam = request.GET['stage']
        keywordParam = request.GET['keyword']

        # data, status = eng_svc.get_dashboard_expanded_engs
        # (stageParam, keywordParam, 0, sys.maxint, user)
        vfs, deployment_targets = eng_svc.get_expanded_engs_for_export(
            stageParam, keywordParam, user)

        workbook = Workbook(write_only=True)
        # Create 'Validation Details' sheet and fill it up with required data:
        validationWorkSheet = workbook.create_sheet()
        validationWorkSheet.title = 'Validation Details'
        headlines = [
            WriteOnlyCell(validationWorkSheet, value=u"EId"),
            WriteOnlyCell(validationWorkSheet, value=u"Engagement"),
            WriteOnlyCell(validationWorkSheet, value=u"Reviewer"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Peer reviewer"),
            WriteOnlyCell(validationWorkSheet, value=u"VFC"),
            WriteOnlyCell(validationWorkSheet, value=u"VFC #"),
            WriteOnlyCell(validationWorkSheet, value=u"Started"),
            WriteOnlyCell(validationWorkSheet, value=u"Vendor"),
            WriteOnlyCell(validationWorkSheet, value=u"AIC Version"),
            WriteOnlyCell(
                validationWorkSheet, value=u"ECOMP Release"),
            WriteOnlyCell(validationWorkSheet, value=u"Validate"),
            WriteOnlyCell(validationWorkSheet, value=u"Completed"),
            WriteOnlyCell(validationWorkSheet, value=u"Stage"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Heat Pre-validated"),
            WriteOnlyCell(validationWorkSheet, value=u"Image Scan"),
            WriteOnlyCell(
                validationWorkSheet, value=u"AIC Instantiated"),
            WriteOnlyCell(
                validationWorkSheet, value=u"ASDC Onboarded"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Overall Progress in %"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Target Completion Date"),
            WriteOnlyCell(validationWorkSheet, value=u"Status")]
        for headline in headlines:
            headline.font = Font(name='Courier', size=16, bold=True)
        validationWorkSheet.append(headlines)

        for vf in vfs:
            validationWorkSheet.append(
                [smart_str(vf["engagement__engagement_manual_id"]),
                 smart_str(vf["vf__name"]),
                 smart_str(
                    vf["vf_engagement__reviewer"]),
                 smart_str(
                    vf["vf_engagement__peer_reviewer"]),
                 smart_str(vf["vfcs"]),
                 smart_str(vf["vfcs__number"]),
                 smart_str(
                    vf["engagement__started_state_time"]),
                 smart_str(vf["vendor__name"]),
                 smart_str(
                    vf["deployment_target__version"]),
                 smart_str(vf["ecomp_release__name"]),
                 smart_str(
                    vf["engagement__validated_time"]),
                 smart_str(
                    vf["engagement__completed_time"]),
                 smart_str(
                    vf["engagement__engagement_stage"]),
                 smart_str(
                    vf["engagement__heat_validated_time"]),
                 smart_str(
                    vf["engagement__image_scan_time"]),
                 smart_str(
                    vf["engagement__aic_instantiation_time"]),
                 smart_str(
                    vf["engagement__asdc_onboarding_time"]),
                 smart_str(vf["engagement__progress"]),
                 smart_str(
                    vf["engagement__target_completion_date"]),
                 smart_str(
                    vf["engagement__latest_status"])
                 ])

        # Create 'Overview' sheet and fill it up with required data:
        overviewWorkSheet = workbook.create_sheet()
        overviewWorkSheet.title = 'Overview'
        headlines = [
            WriteOnlyCell(validationWorkSheet, value=u"AIC/ECOMP"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Active Count of Engagement"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Sum of Nr of VFs"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Intake Count of Engagement"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Sum of Nr of VFs"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Completed Count of Engagement"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Sum of Nr of VFs"),
            WriteOnlyCell(
                validationWorkSheet, value=u"Total Count of Engagement"),
            WriteOnlyCell(validationWorkSheet,
                          value=u"Total Sum of Nr of VFs")]
        for headline in headlines:
            headline.font = Font(name='Courier', size=16, bold=True)
        overviewWorkSheet.append(headlines)

        for deployment_target in deployment_targets:
            overviewWorkSheet.append(deployment_target)

        # We are using HttpResponse and not
        # Rest Response since we couldnt find
        # support for content diposition
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.\
            spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=D2ICE.xlsx'

        workbook.save(response)

        return response


@classDecorator([logFuncEntry])
class GetEngByUser(VvpApiView):

    @auth(Permissions.eng_membership)
    def get(self, request):
        user = request_data_mgr.get_user()
        vf_list = []

        engStageList = [
            EngagementStage.Intake.name, EngagementStage.Active.name,
            EngagementStage.Validated.name, EngagementStage.Completed.name]

        # @UndefinedVariable
        if (user.role.name == Roles.admin.name or
                user.role.name == Roles.admin_ro.name):
            vf_list = VF.objects.filter(
                engagement__engagement_stage__in=engStageList)\
                .distinct().order_by('engagement__engagement_manual_id')\
                .annotate(
                    engagement_manual_id=F('engagement__engagement_manual_id'),
                    engagement_uuid=F('engagement__uuid'),
                    vf_name=F('name'),
                    peer_reviewer__uuid=F('engagement__peer_reviewer__uuid'),
                    engagement_stage=F('engagement__engagement_stage'),
            )\
                .values(
                    'uuid',
                    'vf_name',
                    'engagement_manual_id',
                    'engagement_uuid',
                    'peer_reviewer__uuid',
                    'engagement_stage',
            )
        else:
            vf_list = VF.objects.filter(
                engagement__engagement_stage__in=engStageList).\
                filter(Q(engagement__engagement_team__uuid=user.uuid))\
                .distinct().order_by('engagement__engagement_manual_id')\
                .annotate(
                    engagement_manual_id=F('engagement__engagement_manual_id'),
                    engagement_uuid=F('engagement__uuid'),
                    vf_name=F('name'),
                    peer_reviewer__uuid=F('engagement__peer_reviewer__uuid'),
                    reviewer__uuid=F('engagement__reviewer__uuid'),
                    engagement_stage=F('engagement__engagement_stage'),
            )\
                .values(
                    'uuid',
                    'vf_name',
                    'engagement_manual_id',
                    'engagement_uuid',
                    'peer_reviewer__uuid',
                    'reviewer__uuid',
                    'engagement_stage',
            )

        vf_final_array = []
        for vf in vf_list:
            vfc_list = VFC.objects.filter(Q(vf=vf['uuid'])).values('name')
            vf['vfc'] = []
            for vfc in vfc_list:
                vf['vfc'].append(vfc['name'])

            vf['vfc'] = ', '.join(vf['vfc'])
            vf_final_array.append(vf)

        return Response(vf_final_array)


@classDecorator([logFuncEntry])
class SingleEngByUser(VvpApiView):

    @auth(Permissions.eng_membership)
    def get(self, request, eng_uuid):
        eng_uuid = request_data_mgr.get_eng_uuid()
        user = request_data_mgr.get_user()

        engagement = None

        # @UndefinedVariable
        if (user.role.name == Roles.admin.name or
                user.role.name == Roles.admin_ro.name):
            engagement = Engagement.objects.get(uuid=eng_uuid)
        else:
            try:
                engagement = Engagement.objects.get(
                    peer_reviewer=user, uuid=eng_uuid)
            except Engagement.DoesNotExist:
                try:
                    engagement = Engagement.objects.get(
                        engagement_team__uuid=user.uuid, uuid=eng_uuid)
                except BaseException:
                    msg = "Eng for the User with uuid " + \
                        user.uuid + " doesn't exist."
                    self.logger.error(msg)
                    msg = "Couldn't provide VF, internal server error"
                    return Response(msg, status=HTTP_500_INTERNAL_SERVER_ERROR)
        vfObj = VF.objects.get(engagement__uuid=engagement.uuid)
        eng_svc.update_or_insert_to_recent_engagements(
            user.uuid, vfObj, RecentEngagementActionType.
            NAVIGATED_INTO_ENGAGEMENT.name)

        vfList = VF.objects.filter(engagement__uuid=engagement.uuid)
        formated_vf_list = VFModelSerializer(vfList, many=True).data
        formated_vf = formated_vf_list[0]
        if vfObj.git_repo_url != str(-1):
            formated_vf['files'] = GitManager().getRepoAssociatedFilesForUser(
                eng_uuid)
        else:
            formated_vf['files'] = []

        return Response(formated_vf)

    @auth(Permissions.edit_stage)
    def put(self, request, eng_uuid, stage):
        sts = HTTP_202_ACCEPTED
        if not eng_uuid or not stage:
            msg = "One of the Engagement's input parameters is missing"
            self.logger.error(msg)
            return Response(msg, status=HTTP_400_BAD_REQUEST)
        eng_uuid = request_data_mgr.get_eng_uuid()
        msg = eng_svc.set_engagement_stage(eng_uuid, stage)
        return Response(msg, status=sts)


@classDecorator([logFuncEntry])
class ArchiveEngagement(VvpApiView):

    @auth(Permissions.archive_engagement)
    def put(self, request, eng_uuid):
        data = request.data
        eng_uuid = request_data_mgr.get_eng_uuid()
        msg = eng_svc.archive_engagement(eng_uuid, data['reason'])
        return Response(msg)


@classDecorator([logFuncEntry])
class StarredEngagements(VvpApiView):

    @auth(Permissions.star_an_engagement)
    def get(self, request):
        user = request_data_mgr.get_user()
        if (not user):
            msg = "User with uuid " + user.uuid + \
                " doesn't exist. Can't fetch their engagements"
            self.logger.error(logEncoding(msg))
            msg = "You are not registered as a user, please sign\
             up in order to perform this action"
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        vf_list = eng_svc.vf_retreiver(user, True)
        if not vf_list:
            msg = "Action succeeded, no starred VFs were found."
            self.logger.debug(msg)
            return Response(msg, status=HTTP_204_NO_CONTENT)

        vfListData = []

        for vf_data in vf_list:

            # @UndefinedVariable
            if (
                    user.role.name == Roles.el.name or
                    user.role.name == Roles.admin.name or
                    user.role.name == Roles.admin_ro.name):
                if (vf_data['engagement__reviewer__uuid'] == user.uuid):
                    checklist_lists = Checklist.objects.filter(
                        Q(engagement__uuid=vf_data['engagement__uuid']),
                        Q(engagement__reviewer=user), ~Q(
                            state=CheckListState.archive.name)).values(
                                'uuid', 'name', 'state', 'owner__uuid')
                # @UndefinedVariable
                elif (user.role.name == Roles.admin.name or
                      user.role.name == Roles.admin_ro.name):
                    checklist_lists = Checklist.objects.filter(
                        Q(engagement__uuid=vf_data['engagement__uuid']),
                        ~Q(state=CheckListState.archive.name)).values(
                        'uuid', 'name', 'state', 'owner__uuid')
                else:
                    checklist_lists = Checklist.objects.filter(
                        Q(engagement__uuid=vf_data['engagement__uuid']),
                        Q(owner=user), ~Q(
                            state=CheckListState.archive.name)).values(
                                'uuid', 'name', 'state', 'owner__uuid')
                vf_data['checklists'] = checklist_lists
            else:
                vf_data['checklists'] = None

            vfListData.append(vf_data)

        return Response(vfListData)

    @auth(Permissions.star_an_engagement)
    def put(self, request, **kwargs):
        user = request_data_mgr.get_user()
        data = request.data
        sts = HTTP_200_OK
        if ('engagement_uuid' not in data or not data['engagement_uuid']):
            msg = "starred engagement uuid parameter is missing"
            self.logger.error(msg)
            msg = "Action was failed due to bad request."
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        eng_uuid = data['engagement_uuid']
        if (not user):
            msg = "User with uuid " + user.uuid + \
                " doesn't exist. Can't fetch their engagements"
            self.logger.error(logEncoding(msg))
            msg = "You are not registered as a user, please sign up in order \
            to perform this action"
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        msg = eng_svc.star_an_engagement(user, eng_uuid)

        return Response(msg, status=sts)


@classDecorator([logFuncEntry])
class GetRecentEngagements(VvpApiView):

    @auth(Permissions.eng_membership)
    def get(self, request, format=None, **kwargs):
        user = request_data_mgr.get_user()
        if (not user):
            msg = "User with uuid " + user.uuid + \
                " doesn't exist. Can't fetch their engagements"
            self.logger.error(logEncoding(msg))
            msg = "You are not registered as a user, please sign up in order \
            to perform this action"
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        stared_list = eng_svc.vf_retreiver(user, True)

        vf_list = eng_svc.vf_retreiver(user, False, True)

        if not vf_list:
            msg = "Action succeeded, no recent VFs were found."
            self.logger.debug(msg)
            return Response(msg, status=HTTP_204_NO_CONTENT)

        vfListData = []
        recentList = 20 + stared_list.count()
        for idx, vf_data in enumerate(vf_list):
            if (idx == recentList):
                break
            # @UndefinedVariable
            if (user.role.name == Roles.el.name or
                user.role.name == Roles.admin.name or
                    user.role.name == Roles.admin_ro.name):
                if (vf_data['vf__engagement__reviewer__uuid'] == user.uuid):
                    checklist_lists = Checklist.objects.filter(
                        Q(engagement__uuid=vf_data['vf__engagement__uuid']),
                        Q(engagement__reviewer=user), ~Q(
                            state=CheckListState.archive.name)).values(
                                'uuid', 'name', 'state', 'owner__uuid')
                # @UndefinedVariable
                elif (user.role.name == Roles.admin.name or
                      user.role.name == Roles.admin_ro.name):
                    checklist_lists = Checklist.objects.filter(
                        Q(engagement__uuid=vf_data['vf__engagement__uuid']),
                        ~Q(state=CheckListState.archive.name)).values(
                            'uuid', 'name', 'state', 'owner__uuid')
                else:
                    checklist_lists = Checklist.objects.filter(
                        Q(engagement__uuid=vf_data['vf__engagement__uuid']),
                        Q(owner=user), ~Q(state=CheckListState.archive.name))\
                        .values('uuid', 'name', 'state', 'owner__uuid')
                vf_data['checklists'] = checklist_lists
            else:
                vf_data['checklists'] = None
            vfListData.append(vf_data)
        return Response(vfListData)


@classDecorator([logFuncEntry])
class EngagementProgressBar(VvpApiView):

    @auth(Permissions.edit_progress_bar)
    def put(self, request, eng_uuid):
        data = request.data
        msg = "OK"

        if ('progress' not in data or not data['progress'] or
                data['progress'] == ''):
            msg = "progress parameter is missing or empty"
            self.logger.error(msg)
            msg = "Action has failed due to bad request."
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        eng_svc.set_progress_for_engagement(progress=data['progress'])
        return Response(msg, status=HTTP_202_ACCEPTED)

    @auth(Permissions.get_progress_bar)
    def post(self, request, eng_uuid):
        data = request.data
        eng = self.get_entity(Engagement, request_data_mgr.get_eng_uuid())
        msg = "OK"

        if ('target_date' not in data or not data['target_date'] or
                data['target_date'] == ''):
            msg = "target_date parameter is missing or empty"
            self.logger.error(msg)
            msg = "Action has failed due to bad request."
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        target_date = data['target_date']
        date = parse_date(target_date)
        eng.target_completion_date = date
        eng.save()
        return Response(msg)


@classDecorator([logFuncEntry])
class ChangeTargetLabEntryDate(VvpApiView):

    @auth(Permissions.change_lab_entry)
    def post(self, request, eng_uuid):
        data = request.data
        eng = self.get_entity(Engagement, request_data_mgr.get_eng_uuid())
        msg = "OK"
        vf = VF.objects.get(engagement__uuid=eng.uuid)
        if ('target_date' not in data or not data['target_date']
                or data['target_date'] == ''):
            msg = "target_date parameter is missing or empty"
            self.logger.error(msg)
            msg = "Action has failed due to bad request."
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        target_date = data['target_date']
        date = parse_date(target_date)
        vf.target_lab_entry_date = date
        vf.save()
        return Response(msg)


@classDecorator([logFuncEntry, ])
class Status(VvpApiView):

    @auth(Permissions.get_engagement_status)
    def get(self, request, eng_uuid):
        eng_uuid = request_data_mgr.get_eng_uuid()
        engagement = Engagement.objects.get(uuid=eng_uuid)
        latest_status = EngagementStatus.objects.filter(
            engagement=engagement).distinct().order_by('-update_time')[:1]
        if (latest_status.count() > 0):
            latest_status = latest_status[0]
            latest_status = EngagementStatusModelSerializer(latest_status).data
        else:
            latest_status = False
        return Response(latest_status)

    @auth(Permissions.put_engagement_status)
    def put(self, request, eng_uuid):
        user = request_data_mgr.get_user()
        data = request.data

        eng_uuid = request_data_mgr.get_eng_uuid()
        eng_status_uuid = data['eng_status_uuid']
        description = data['description']

        if not description:
            msg = "Not description sent"
            self.logger.error(msg)
            msg = "You are not registered as a user, please sign \
            up in order to perform this action"
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        # @UndefinedVariable
        if (user.role.name != Roles.admin.name and
                user.role.name != Roles.el.name):
            msg = "User not authorized"
            self.logger.error(msg)
            msg = "Internal error."
            return Response(msg, status=HTTP_401_UNAUTHORIZED)

        engagement = Engagement.objects.get(uuid=eng_uuid)
        status = EngagementStatus.objects.get(uuid=eng_status_uuid)
        eng_svc.update_engagement_status(
            user, description, eng_status_uuid, engagement)

        status = EngagementStatus.objects.get(uuid=eng_status_uuid)
        status = EngagementStatusModelSerializer(status).data

        return Response(status)

    @auth(Permissions.put_engagement_status)
    def post(self, request, eng_uuid):
        user = request_data_mgr.get_user()
        eng_uuid = request_data_mgr.get_eng_uuid()
        data = request.data

        description = data['description']
        if not description:
            msg = "No description sent"
            self.logger.error(msg)
            msg = "No description sent"
            return Response(msg, status=HTTP_400_BAD_REQUEST)

        # @UndefinedVariable
        if (user.role.name != Roles.admin.name and
                user.role.name != Roles.el.name):
            msg = "User not authorized"
            self.logger.error(msg)
            msg = "Internal error."
            return Response(msg, status=HTTP_401_UNAUTHORIZED)
        engagement = Engagement.objects.get(uuid=eng_uuid)
        created_eng_staus = eng_svc.insert_engagement_status(
            user, description, engagement)
        created_eng_staus = EngagementStatusModelSerializer(
            created_eng_staus).data
        return Response(created_eng_staus)


@classDecorator([logFuncEntry])
class EngagementOps(VvpApiView):

    @auth(Permissions.update_engagement)
    def put(self, request, eng_uuid):
        data = request.data
        engagement_dict = data['engagement']
        engagement = eng_svc.update_engagement(engagement_dict)

        if engagement is None:
            return Response(status=HTTP_400_BAD_REQUEST)
        else:
            if 'status' in data and data['status']:
                user = request_data_mgr.get_user()
                created_eng_status = eng_svc.insert_engagement_status(
                    user, data['status'], engagement)
                return Response(
                    EngagementStatusModelSerializer(created_eng_status).data)
            else:
                return Response()


@classDecorator([logFuncEntry])
class EngagementTeamUsers(VvpApiView):

    @auth(Permissions.remove_from_engagement_team)
    def put(self, request):
        eng_uuid = request_data_mgr.get_eng_uuid()
        user = request_data_mgr.get_user()
        data = request.data
        if (data['user_uuid']):
            requested_user_uuid = data['user_uuid']
        if (eng_uuid and user and data['user_uuid']):
            eng_svc.remove_user_from_engagement_team(
                eng_uuid, user, requested_user_uuid)
        return Response(status=HTTP_204_NO_CONTENT)


@classDecorator([logFuncEntry])
class EngagementReviewer(VvpApiView):

    @auth(Permissions.update_engagement_reviewers)
    def put(self, request, eng_uuid):
        data = request.data
        eng_uuid = request_data_mgr.get_eng_uuid()
        reviewer = eng_svc.set_engagement_reviewer(eng_uuid, data['reviewer'])

        if reviewer is None:
            return Response(status=HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            return Response(reviewer)


@classDecorator([logFuncEntry])
class EngagementPeerReviewer(VvpApiView):

    @auth(Permissions.update_engagement_reviewers)
    def put(self, request, eng_uuid):
        data = request.data
        eng_uuid = request_data_mgr.get_eng_uuid()
        peer_reviewer = eng_svc.set_engagement_peer_reviewer(
            eng_uuid, data['peerreviewer'])

        if peer_reviewer is None:
            return Response(status=HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            return Response(peer_reviewer)


@classDecorator([logFuncEntry])
class SwitchEngagementReviewers(VvpApiView):

    @auth(Permissions.update_engagement_reviewers)
    def put(self, request, eng_uuid):
        data = request.data
        eng_uuid = request_data_mgr.get_eng_uuid()
        resp = eng_svc.switch_engagement_reviewers(
            eng_uuid, data['reviewer'], data['peerreviewer'])
        if resp is None:
            return Response(status=HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            return Response(resp)
